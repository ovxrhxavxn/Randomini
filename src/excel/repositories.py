import io
from typing import Protocol

import openpyxl


class AbstractExcelRepository(Protocol):

    def get_values(self, file_bytes: bytes) -> list[dict]:
        pass


class ExcelRepository:

    def get_values(self, file_bytes: bytes) -> list[dict]:
        """
        Парсит XLSX из байтов и возвращает список словарей.
        Использует openpyxl в режиме read-only для производительности.
        """
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheet = workbook.active

        # Предполагаем, что первая строка - это заголовок
        headers = [cell.value for cell in sheet[1]]
        data = []

        # iter_rows() - эффективный итератор
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))

        return data